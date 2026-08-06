from __future__ import annotations

import re
import warnings
from io import BytesIO
from typing import Any, BinaryIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADERS = ["과목", "학년", "학기", "번호", "성명", "세부능력 및 특기사항"]


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalized_label(value: Any) -> str:
    return re.sub(r"\s+", "", clean_text(value))


def integer_value(value: Any) -> int | None:
    if isinstance(value, bool) or value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return int(number) if number.is_integer() else None


def is_column_header(values: tuple[Any, ...]) -> bool:
    return (
        normalized_label(values[0]) == "과목"
        and normalized_label(values[3]) == "번호"
        and normalized_label(values[4]) == "성명"
    )


def is_page_footer(values: tuple[Any, ...]) -> bool:
    current_page = integer_value(values[7])
    total_pages = integer_value(values[9])
    return current_page is not None and clean_text(values[8]) == "/" and total_pages is not None


def extract_records(source: BinaryIO) -> tuple[list[list[Any]], int, int]:
    """조회형 Excel을 읽어 정규화 레코드, 원본 행 수, 병합 수를 반환한다."""
    source.seek(0)
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Workbook contains no default style")
        # 원본의 XML dimension이 A1로 잘못 기록될 수 있어 read_only 모드는 쓰지 않는다.
        workbook = load_workbook(source, data_only=True, read_only=False)

    sheet = workbook[workbook.sheetnames[0]]
    records: list[list[Any]] = []
    current_subject: str | None = None
    current_grade: int | None = None
    current_semester: int | None = None
    page_break_since_last_data = False
    physical_data_rows = 0
    merged_fragments = 0

    try:
        for row in sheet.iter_rows(min_col=1, max_col=11, values_only=True):
            values = tuple(row)
            if is_page_footer(values):
                page_break_since_last_data = True
                continue
            if is_column_header(values):
                continue

            number = integer_value(values[3])
            name = clean_text(values[4])
            detail = "" if values[5] is None else str(values[5])
            if number is None or not name or not detail.strip():
                continue

            physical_data_rows += 1
            if clean_text(values[0]):
                current_subject = clean_text(values[0])
            if values[1] is not None:
                current_grade = integer_value(values[1])
            if values[2] is not None:
                current_semester = integer_value(values[2])

            if current_subject is None or current_grade is None or current_semester is None:
                raise ValueError(
                    f"'{name}'(번호 {number}) 행의 과목·학년·학기 정보를 복원할 수 없습니다."
                )

            key = (current_subject, current_grade, current_semester, number, name)
            previous_key = tuple(records[-1][:5]) if records else None
            if records and page_break_since_last_data and previous_key == key:
                records[-1][5] += detail
                merged_fragments += 1
            else:
                records.append([*key, detail])

            page_break_since_last_data = False
    finally:
        workbook.close()

    return records, physical_data_rows, merged_fragments


def build_workbook(records: list[list[Any]]) -> BytesIO:
    output = Workbook()
    sheet = output.active
    sheet.title = "정리"
    sheet.append(HEADERS)
    for record in records:
        sheet.append(record)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(name="맑은 고딕", size=11, color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 8
    sheet.column_dimensions["C"].width = 8
    sheet.column_dimensions["D"].width = 8
    sheet.column_dimensions["E"].width = 12
    sheet.column_dimensions["F"].width = 100

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(vertical="top")
        row[5].alignment = Alignment(vertical="top", wrap_text=True)

    table = Table(displayName="StudentRecords", ref=f"A1:F{sheet.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    result = BytesIO()
    output.save(result)
    output.close()
    result.seek(0)
    return result
